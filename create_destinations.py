from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

wb = Workbook()

# ── Sheet 1: Destinations ──────────────────────────────────────────────────
ws = wb.active
ws.title = "Destinations"

HEADER_FILL   = PatternFill("solid", start_color="0F1C3F")
SUBHEAD_FILL  = PatternFill("solid", start_color="1A3A6B")
ALT_ROW_FILL  = PatternFill("solid", start_color="EEF2F9")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(name="Calibri", bold=True, color="C8D8FF", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_BODY     = Font(name="Calibri", bold=True, size=10)
NOTE_FONT     = Font(name="Calibri", italic=True, color="888888", size=9)

THIN  = Side(style="thin",   color="CCCCCC")
MED   = Side(style="medium", color="0F1C3F")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Title row
ws.merge_cells("A1:O1")
title_cell = ws["A1"]
title_cell.value = "GALAXIUM TRAVELS — Destination Content Sheet"
title_cell.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
title_cell.fill  = PatternFill("solid", start_color="0A0F2C")
title_cell.alignment = CENTER
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:O2")
sub = ws["A2"]
sub.value = "Maintained by the Content Team · Do not edit columns A–C without consulting Engineering · Last updated: 2099-01-15"
sub.font  = NOTE_FONT
sub.fill  = PatternFill("solid", start_color="D6E4FF")
sub.alignment = LEFT
ws.row_dimensions[2].height = 18

# Column headers + widths
columns = [
    ("Slug",              16),
    ("Display_Name",      18),
    ("Tagline",           38),
    ("Body_Type",         14),
    ("Gravity_g",         13),
    ("Distance_AU",       14),
    ("Transit_Days",      14),
    ("Atmosphere",        28),
    ("Surface_Temp_Min",  16),
    ("Surface_Temp_Max",  16),
    ("Hazard_Level",      14),
    ("Hazards",           40),
    ("Fun_Fact",          45),
    ("Emoji",              8),
    ("Active",            10),
]

for col_idx, (header, width) in enumerate(columns, start=1):
    col_letter = ws.cell(row=3, column=col_idx).column_letter
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font  = HEADER_FONT
    cell.fill  = HEADER_FILL
    cell.alignment = CENTER
    cell.border = CELL_BORDER
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 28

# Units sub-header
units = [
    "url-safe key", "human name", "marketing one-liner", "classification",
    "x Earth (1.0)", "AU", "Earth days", "composition",
    "deg C", "deg C", "1=safe 5=extreme", "comma list", "one sentence", "icon", "show on site"
]
for col_idx, unit in enumerate(units, start=1):
    cell = ws.cell(row=4, column=col_idx, value=unit)
    cell.font  = SUBHEAD_FONT
    cell.fill  = SUBHEAD_FILL
    cell.alignment = CENTER
    cell.border = CELL_BORDER
ws.row_dimensions[4].height = 20

# Destination data
destinations = [
    ("earth",   "Earth",       "Where it all began — and where your journey ends.",                    "Planet",       1.00,  0.00,   0,   "Nitrogen-oxygen",                         -88, 58,   1, "Weather variability",                                "Earth is the only known body in the universe confirmed to harbour life — for now.",                                                         "🌍", True),
    ("moon",    "The Moon",    "One giant leap, now a two-hour commute.",                               "Moon",         0.17,  0.0026, 0.17,"Negligible (exosphere)",                  -173, 127,  2, "Micrometeorites, radiation, vacuum",                 "The Moon is slowly drifting away from Earth at roughly 3.8 cm per year.",                                                                   "🌕", True),
    ("mars",    "Mars",        "Red, rusty, and ready for settlers.",                                   "Planet",       0.38,  1.52,   8,   "Thin CO2 (0.6% Earth pressure)",          -125, 20,   3, "Dust storms, radiation, low pressure",               "Olympus Mons on Mars is the tallest volcano in the solar system at ~21 km.",                                                                "🔴", True),
    ("venus",   "Venus",       "Glamorous from afar, scorching up close.",                              "Planet",       0.91,  0.72,   5,   "Dense CO2 with sulfuric acid clouds",      462,  465,  5, "Acid rain, extreme heat, crushing pressure",         "Venus rotates so slowly that a day on Venus is longer than its year.",                                                                      "🟡", True),
    ("jupiter", "Jupiter",     "King of planets — bring a pressure suit.",                              "Planet",       2.53,  5.20,   30,  "Hydrogen and helium — no solid surface",  -145, None, 4, "Radiation belts, no solid surface, storms",          "Jupiter's Great Red Spot is a storm that has raged for at least 350 years.",                                                                "🟠", True),
    ("europa",  "Europa",      "An ocean world hiding secrets under the ice.",                          "Moon",         0.13,  5.20,   32,  "Thin oxygen (trace)",                     -160, -160, 4, "Radiation, ice fissures, tidal flexing",             "Europa is believed to harbour a liquid water ocean beneath its icy crust — a prime candidate for extraterrestrial life.",                   "🧊", True),
    ("pluto",   "Pluto",       "Demoted from planet, elevated to legend.",                              "Dwarf Planet", 0.063, 39.48,  330, "Thin nitrogen with methane traces",       -233, -218, 5, "Extreme cold, radiation, micro-gravity",             "Pluto has a heart-shaped nitrogen-ice plain called Tombaugh Regio, visible from orbit.",                                                    "💜", True),
]

for row_idx, dest in enumerate(destinations, start=5):
    fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
    for col_idx, value in enumerate(dest, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font   = BODY_FONT
        cell.fill   = fill
        cell.border = CELL_BORDER
        cell.alignment = LEFT
    for col in [1, 4, 5, 6, 7, 9, 10, 11, 14, 15]:
        ws.cell(row=row_idx, column=col).alignment = CENTER
    ws.cell(row=row_idx, column=1).font = BOLD_BODY
    ws.row_dimensions[row_idx].height = 40

ws.freeze_panes = "A5"

# ── Sheet 2: Hazard Legend ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Hazard Legend")
legend_data = [
    ("Hazard Level", "Label",     "Description",                                           "Hex Colour"),
    (1,              "Safe",      "Standard tourist destination. No special gear needed.",  "#4CAF50"),
    (2,              "Low Risk",  "Minor hazards. Basic precautions advised.",              "#8BC34A"),
    (3,              "Moderate",  "Significant environmental hazards. Guided tours only.",  "#FFC107"),
    (4,              "High Risk", "Dangerous conditions. Specialised suits required.",      "#FF5722"),
    (5,              "Extreme",   "Life-threatening. Research missions only.",              "#D32F2F"),
]
hazard_hex = {1: "4CAF50", 2: "8BC34A", 3: "FFC107", 4: "FF5722", 5: "D32F2F"}

for r_idx, row in enumerate(legend_data, start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.border = CELL_BORDER
        cell.alignment = LEFT
        if r_idx == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = BODY_FONT
            if c_idx == 1:
                cell.fill = PatternFill("solid", start_color=hazard_hex[val])
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

for col, width in zip(["A","B","C","D"], [14, 12, 52, 14]):
    ws2.column_dimensions[col].width = width

# ── Sheet 3: Engineering Notes ─────────────────────────────────────────────
ws3 = wb.create_sheet("Engineering Notes")
notes = [
    ("Field",             "TypeScript Type",                    "Notes"),
    ("slug",              "string",                             "Used as URL key: /destinations/:slug"),
    ("display_name",      "string",                             "Human-readable name shown in headings"),
    ("tagline",           "string",                             "Short marketing line, max ~60 chars"),
    ("body_type",         "'Planet' | 'Moon' | 'Dwarf Planet'", "Badge label in the UI"),
    ("gravity_g",         "number",                             "Relative to Earth (1.0)"),
    ("distance_au",       "number",                             "Average distance from Earth in AU"),
    ("transit_days",      "number",                             "One-way trip duration in Earth days"),
    ("atmosphere",        "string",                             "Brief composition description"),
    ("surface_temp_min",  "number | null",                      "deg C — null if no solid surface (e.g. Jupiter)"),
    ("surface_temp_max",  "number | null",                      "deg C — null if no solid surface"),
    ("hazard_level",      "1 | 2 | 3 | 4 | 5",                 "1=safe, 5=extreme — drives badge colour (see Hazard Legend sheet)"),
    ("hazards",           "string[]",                           "Split on ', ' when importing to TypeScript"),
    ("fun_fact",          "string",                             "Single sentence shown in destination card footer"),
    ("emoji",             "string",                             "Decorative only; not used as accessible label"),
    ("active",            "boolean",                            "FALSE rows are excluded from the UI entirely"),
]
for r_idx, row in enumerate(notes, start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.border = CELL_BORDER
        cell.alignment = LEFT
        if r_idx == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = BODY_FONT

for col, width in zip(["A","B","C"], [22, 36, 58]):
    ws3.column_dimensions[col].width = width

wb.save("destinations_content.xlsx")
print("Done: destinations_content.xlsx")
